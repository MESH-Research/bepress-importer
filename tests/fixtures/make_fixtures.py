"""Generate the binary Excel fixtures (mini.xls, mini.xlsx) committed to the repo.

Run manually with `uv run python tests/fixtures/make_fixtures.py` when the
fixture content needs to change; never runs in CI. The generated files must
stay logically identical to mini.csv so reader-equivalence tests hold.
"""

from datetime import date, datetime
from pathlib import Path

HERE = Path(__file__).parent

COLUMNS = ["title", "publication_date", "context_key", "price", "empty_col", "note"]
ROWS = [
    ["First Item", date(2015, 4, 1), 2404130, 3.5, "", "padded"],
    ['Second, "quoted"', date(1998, 1, 1), 2404131, 12, "", "unicode: séance—ok"],
]


def make_xls() -> None:
    import xlwt

    book = xlwt.Workbook()
    sheet = book.add_sheet("mini")
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    for col, header in enumerate(COLUMNS):
        sheet.write(0, col, header)
    for r, row in enumerate(ROWS, start=1):
        for c, value in enumerate(row):
            if isinstance(value, (date, datetime)):
                sheet.write(r, c, value, date_style)
            elif c == 5 and r == 1:
                sheet.write(r, c, "  padded  ")  # exercises whitespace stripping
            else:
                sheet.write(r, c, value)
    # a fully empty trailing row must be skipped by readers
    sheet.write(4, 0, "")
    book.save(str(HERE / "mini.xls"))


def make_xlsx() -> None:
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "mini"
    sheet.append(COLUMNS)
    for r, row in enumerate(ROWS, start=2):
        for c, value in enumerate(row, start=1):
            cell = sheet.cell(row=r, column=c)
            if c == 6 and r == 2:
                cell.value = "  padded  "
            else:
                cell.value = value
            if isinstance(value, (date, datetime)):
                cell.number_format = "YYYY-MM-DD"
    book.save(str(HERE / "mini.xlsx"))


if __name__ == "__main__":
    make_xls()
    make_xlsx()
    print("wrote", HERE / "mini.xls", "and", HERE / "mini.xlsx")
