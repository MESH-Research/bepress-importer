"""Read .xls, .xlsx and .csv exports into normalized tables.

Every cell value is normalized to a string: Excel date serials become ISO
dates, integral floats lose their trailing ".0", empty cells become "".
"""

from __future__ import annotations

import csv
import datetime as dt
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class Table:
    """One sheet: a name, ordered column headers, and rows as column→value dicts."""

    name: str
    columns: tuple[str, ...]
    rows: tuple[dict[str, str], ...]


@dataclass(frozen=True)
class Workbook:
    """An ordered collection of tables read from one input file."""

    tables: tuple[Table, ...]


def read_workbook(path: str | Path) -> Workbook:
    """Read a .xls, .xlsx or .csv file into a Workbook of normalized tables."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Unsupported input format {suffix!r} (expected .xls, .xlsx or .csv)")


def _dedupe_headers(headers: list[str]) -> tuple[str, ...]:
    seen: dict[str, int] = {}
    result = []
    for header in headers:
        count = seen.get(header, 0) + 1
        seen[header] = count
        result.append(header if count == 1 else f"{header}_{count}")
    return tuple(result)


def _build_table(name: str, headers: list[str], raw_rows: list[list[str]]) -> Table:
    columns = _dedupe_headers([h.strip() for h in headers])
    rows = []
    for raw in raw_rows:
        values = [v.strip() for v in raw]
        # pad or truncate to the header width
        values = (values + [""] * len(columns))[: len(columns)]
        if all(v == "" for v in values):
            continue
        rows.append(dict(zip(columns, values)))
    return Table(name=name, columns=columns, rows=tuple(rows))


def _read_csv(path: Path) -> Workbook:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
    if not rows:
        return Workbook(tables=(Table(name=path.stem, columns=(), rows=()),))
    return Workbook(tables=(_build_table(path.stem, rows[0], rows[1:]),))


def _number_to_str(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return repr(value)


def _read_xls(path: Path) -> Workbook:
    import xlrd

    book = xlrd.open_workbook(str(path))
    tables = []
    for sheet in book.sheets():
        raw_rows: list[list[str]] = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    parts = xlrd.xldate_as_tuple(cell.value, book.datemode)
                    if parts[3] == parts[4] == parts[5] == 0:
                        row.append(dt.date(*parts[:3]).isoformat())
                    else:
                        # spreadsheet cells are wall-clock times with no timezone
                        row.append(dt.datetime(*parts).isoformat(sep=" "))  # noqa: DTZ001
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    row.append(_number_to_str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append("TRUE" if cell.value else "FALSE")
                elif cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    row.append("")
                else:
                    row.append(str(cell.value))
            raw_rows.append(row)
        if not raw_rows:
            tables.append(Table(name=sheet.name, columns=(), rows=()))
            continue
        tables.append(_build_table(sheet.name, raw_rows[0], raw_rows[1:]))
    return Workbook(tables=tuple(tables))


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return _number_to_str(value)
    return str(value)


def _read_xlsx(path: Path) -> Workbook:
    import openpyxl

    book = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    tables = []
    for sheet in book.worksheets:
        raw_rows = [[_cell_to_str(value) for value in row] for row in sheet.iter_rows(values_only=True)]
        if not raw_rows:
            tables.append(Table(name=sheet.title, columns=(), rows=()))
            continue
        tables.append(_build_table(sheet.title, raw_rows[0], raw_rows[1:]))
    book.close()
    return Workbook(tables=tuple(tables))
